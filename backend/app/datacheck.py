import json
import re
from dataclasses import astuple, dataclass
from datetime import datetime
from itertools import product
from pathlib import Path
from typing import List, Optional, Union

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.base import MonitoringData, read_data, read_table
from app.logger import get_logger

logger = get_logger(__name__)


@dataclass
class ErrDat:
    plot_id: str
    rec_id1: str
    rec_id2: str
    err_type: str


class CheckDataCommon(MonitoringData):
    """
    Check errors in ecosystem monitoring data of the Monitoring Sites 1000 project.

    共通のチェック項目
    """

    __fd = Path(__file__).resolve().parents[0]
    __path_spdict_default = __fd.joinpath("suppl_data", "species_dict.json")
    __path_xy_default = __fd.joinpath("suppl_data", "grid_xy.json")
    __path_trap_default = __fd.joinpath("suppl_data", "trap_list.json")

    def __init__(
        self,
        path_spdict: str = "",
        path_xy: str = "",
        path_trap: str = "",
        *args,
        **kwargs
    ):
        super().__init__(*args, **kwargs)
        self.__prepare()

        if self.data_type in ["treeGBH", "seed"]:
            if not path_spdict:
                path_spdict = str(self.__path_spdict_default)
            with open(path_spdict, "rb") as f:
                self.dict_sp = json.load(f)

        if self.data_type == "treeGBH":
            if not path_xy:
                path_xy = str(self.__path_xy_default)
            with open(path_xy, "rb") as f:
                dict_xy = json.load(f)
            if self.plot_id in dict_xy:
                self.xy_combn = list(product(*dict_xy[self.plot_id].values()))
            else:
                self.xy_combn = []

        if self.data_type in ["litter", "seed"]:
            if not path_trap:
                path_trap = str(self.__path_trap_default)
            with open(path_trap, "rb") as f:
                dict_trap = json.load(f)
            if self.plot_id in dict_trap:
                self.trap_list = dict_trap[self.plot_id]
            else:
                self.trap_list = {}

    def __prepare(self):
        if self.data_type == "treeGBH":
            pat_meas_col = "^gbh[0-9]{2}$"
            pat_except = "^(?<![nd])d(?![d])|^dd$|^NA$|^na$|^na<5|^vi|^vn|^cd|^nd"
            pat_rec_id = "tag_no"
        elif self.data_type == "litter":
            pat_meas_col = "^w_|^wdry_"
            pat_except = "^NA$|^na$|^nd|^-$"
            pat_rec_id = "s_date1"
        elif self.data_type == "seed":
            pat_meas_col = "^number|^wdry"
            pat_except = "^NA$|^na$|^nd|^-$"
            pat_rec_id = "s_date1"
        else:
            raise TypeError("The data is not in the format of Moni-sen data")

        self.meas = self.select(regex=pat_meas_col)
        self.col_meas = np.array([x for x in self.columns if re.match(pat_meas_col, x)])
        na_cols = (self.meas == "NA").all(axis=0)
        self.meas = self.meas[:, ~na_cols]
        self.col_meas = self.col_meas[~na_cols]
        self.meas_orig = self.meas.copy()
        self.rec_id = self.select(pat_rec_id)
        self.pat_except = pat_except

        if self.data_type in ["litter", "seed"]:
            self.trap_id = self.select("trap_id")
        else:
            self.trap_id = None

    def __repr__(self):
        return object.__repr__(self)

    def check_invalid_date(self):
        """
        Check for invalid values on the survey dates.

        調査日に不正な入力値
        """
        date_cols = list(filter(lambda x: re.match("^s_date", x), self.columns))
        date = self.select(regex="^s_date")
        date_orig = date.copy()
        date = np.vectorize(lambda x: re.sub("^NA$|^na$|^nd|^$", "11111111", x))(date)
        valid = np.vectorize(lambda x: isdate(x))(date)
        msg = "{}に不正な入力値 ({})"
        errors = [
            ErrDat(
                self.plot_id,
                self.rec_id[i],
                "",
                msg.format(date_cols[j], date_orig[i, j]),
            )
            for i, j in zip(*np.where(~valid))
        ]
        return errors

    def check_sp_not_in_list(self):
        """
        Check if the species name is on the list.

        種リストにない
        """
        errors = []
        if not self.dict_sp:
            return errors
        splist_obs = np.unique(self.select(regex="^spc$|^spc_japan$"))
        sp_not_in_list = [sp for sp in splist_obs if sp not in self.dict_sp]
        msg = "変則的な種名もしくは標準和名だがリストにない種 ({})"
        if sp_not_in_list:
            for sp in sp_not_in_list:
                if self.data_type == "treeGBH":
                    tag = self.rec_id[np.where(self.select("spc_japan") == sp)]
                    rec_id1 = "; ".join(tag)
                else:
                    rec_id1 = ""
                errors.append(ErrDat(self.plot_id, rec_id1, "", msg.format(sp)))
        return errors

    def check_synonym(self):
        """
        Check if the same species is recorded in two or more Japanese name synonyms.

        同種が2つ以上の名前で入力されている
        """
        errors = []
        if not self.dict_sp:
            return errors
        splist_obs = np.unique(self.select(regex="^spc$|^spc_japan$"))
        sp_in_list = np.array([sp for sp in splist_obs if sp in self.dict_sp])
        name_std = np.array([self.dict_sp[sp]["name_jp_std"] for sp in sp_in_list])
        uniq, cnt = np.unique(name_std, return_counts=True)
        dups = uniq[np.where(cnt > 1)]
        for sp in dups:
            if sp:
                synonyms = sp_in_list[np.where(name_std == sp)]
                msg = "同種が2つの名前で入力 ({})".format("/".join(synonyms))
                errors.append(ErrDat(self.plot_id, "", "", msg))
        return errors

    def check_local_name(self):
        """
        Check for non-standard Japanese names.

        非標準和名
        """
        errors = []
        if not self.dict_sp:
            return errors
        splist_obs = np.unique(self.select(regex="^spc$|^spc_japan$"))
        sp_in_list = np.array([sp for sp in splist_obs if sp in self.dict_sp])
        name_std = np.array([self.dict_sp[sp]["name_jp_std"] for sp in sp_in_list])
        for sp, spstd in zip(sp_in_list, name_std):
            if sp != spstd:
                if spstd and not spstd.endswith(("科", "属", "節", "類")):
                    msg = "{}は非標準和名（{}の別名）".format(sp, spstd)
                    if self.data_type == "treeGBH":
                        tag = self.rec_id[np.where(self.select("spc_japan") == sp)]
                        rec_id1 = "; ".join(tag)
                    else:
                        rec_id1 = ""
                    errors.append(ErrDat(self.plot_id, rec_id1, "", msg))
        return errors

    def check_blank_in_data_cols(self):
        """
        Check for blanks in measurements.

        測定値データの空白
        """
        msg = "測定値データ列に空白セルあり"
        errors = [
            ErrDat(
                self.plot_id,
                self.rec_id[i],
                self.col_meas[j] if self.data_type == "treeGBH" else self.trap_id[i],
                msg,
            )
            for i, j in zip(*np.where(self.meas == ""))
        ]
        return errors

    def check_invalid_values(self):
        """
        Check for invalid inputs in measurements.

        測定値の無効な入力値
        """
        valid = np.vectorize(lambda x: isvalid(x, self.pat_except))(self.meas)
        msg = "{}が無効な入力値 ({})"
        errors = [
            ErrDat(
                self.plot_id,
                self.rec_id[i],
                self.col_meas[j] if self.data_type == "treeGBH" else self.trap_id[i],
                msg.format(self.col_meas[j], self.meas_orig[i, j]),
            )
            for i, j in zip(*np.where(~valid))
        ]
        return errors

    def mask_invalid_values(self):
        """
        Mask invalid values in measurements.

        測定値の無効な入力値をnp.nanに置換
        """
        valid = np.vectorize(lambda x: isvalid(x, self.pat_except))(self.meas)
        for i, j in zip(*np.where(~valid)):
            self.meas[i, j] = np.nan

    def check_positive(self):
        """
        Check if measurement values are positive.

        測定値が正の値かどうかのチェック
        """
        meas_c = np.vectorize(lambda x: isvalid(x, return_value=True))(self.meas)
        meas_c[np.isnan(meas_c)] = 0

        msg = "{}の測定値がマイナス ({})"

        errors = [
            ErrDat(
                self.plot_id,
                self.rec_id[i],
                self.col_meas[j] if self.data_type == "treeGBH" else self.trap_id[i],
                msg.format(self.col_meas[j], meas_c[i, j]),
            )
            for i, j in zip(*np.where(meas_c < 0))
        ]
        return errors


class CheckDataTree(CheckDataCommon):
    """
    Check tree GBH data.

    毎木データのチェック
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def check_tag_dup(self):
        """
        Check for tag number duplication.

        タグ番号の重複
        """
        msg = "タグ番号の重複"
        tag_counts = np.unique(self.rec_id, return_counts=True)
        errors = [
            ErrDat(self.plot_id, tag, "", msg) for tag, n in zip(*tag_counts) if n > 1
        ]
        return errors

    def check_indv_null(self):
        """
        Check for blank or na in individual numbers.

        indv_noが空白またはna
        """
        indv_no = self.select("indv_no")
        indv_no = np.array(["" if i in ["na", "NA"] else i for i in indv_no])
        msg = "indv_noが空白またはna"
        return [
            ErrDat(self.plot_id, self.rec_id[i], "indv_no", msg)
            for i in np.where(indv_no == "")[0]
        ]

    def check_sp_mismatch(self):
        """
        Check for species mismatch on the stems of the same individual.

        同株で樹種が異なる
        """
        errors = []
        indv_no = self.select("indv_no")
        indv_no = np.array(["" if i in ["na", "NA"] else i for i in indv_no])
        indv_counts = np.unique(indv_no, return_counts=True)
        sp = self.select("spc_japan")
        msg = "同株だが樹種が異なる"
        for i in indv_counts[0][np.where(indv_counts[1] > 1)]:
            s = np.where(indv_no == i)
            if len(np.unique(sp[s])) > 1 and i:
                errors.append(
                    ErrDat(
                        self.plot_id, "; ".join(self.rec_id[s]), "/".join(sp[s]), msg
                    )
                )
        return errors

    def check_mesh_xy(self):
        """
        Check for errors in  coordinates: 'mesh_[xy]cord'.

        mesh_[xy]cordのエラー
        """
        errors = []
        if not self.xy_combn:
            return errors

        for i, xy in enumerate(self.select(regex="^mesh_[xy]cord$")):
            tag = self.rec_id[i]
            target = "mesh_xycord={}".format(str(xy))
            try:
                if not tuple(xy.astype(int)) in self.xy_combn:
                    msg = "調査地に存在しないxy座標の組み合わせ"
                    errors.append(ErrDat(self.plot_id, tag, target, msg))
            except ValueError:
                if any([j in ["nd", "na", "NA"] for j in xy]):
                    pass
                elif list(xy).count("") > 0:
                    msg = "mesh_xycordに空白セル"
                    errors.append(ErrDat(self.plot_id, tag, target, msg))
                else:
                    msg = "mesh_xycordの入力値が非数値"
                    errors.append(ErrDat(self.plot_id, tag, target, msg))
        return errors

    def check_stem_xy(self):
        """
        Check for errors in xy coordinates: 'stem_[xy]cord'.

        stem_[xy]cordのエラー
        """
        errors = []
        if not self.xy_combn:
            return errors

        for i, xy in enumerate(self.select(regex="^stem_[xy]cord$")):
            tag = self.rec_id[i]
            target = "stem_xycord={}".format(str(xy))
            try:
                np.array(xy).astype(np.float64)
            except ValueError:
                if any([j in ["nd", "na", "NA"] for j in xy]):
                    pass
                elif list(xy).count("") > 0:
                    msg = "stem_xycordに空白セル"
                    errors.append(ErrDat(self.plot_id, tag, target, msg))
                else:
                    msg = "stem_xycordの入力値が非数値"
                    errors.append(ErrDat(self.plot_id, tag, target, msg))
        return errors

    def replace_dxx_in_gbh(self):
        """
        Convert to 'd' if the value is 'dxx.xx'.

        枯死個体のgbhが'dxx.xx'と入力されている場合は'd'に変換
        """
        pat_dxx = r"(?<![nd])d(?![d])\s?([0-9]+[.]?[0-9]*)"
        match_dxx = np.vectorize(lambda x: find_pattern(x, pat_dxx))(self.meas)
        for i, j in zip(*np.where(match_dxx)):
            if j > 0 and match_dxx[i, j - 1]:
                # 複数年に渡って"dxx.xx"と入力されている場合は、2つ目以降を'na'にする
                self.meas[i, j] = "na"
            else:
                self.meas[i, j] = "d"

    def check_missing(self):
        """
        Check for 'na' values for individuals that were alive at the last census.

        前年まで生存していた個体がnaになっている
        """
        pat_na = r"^na$|^NA$"
        match_na = np.vectorize(lambda x: find_pattern(x, pat_na))(self.meas)
        alive = np.vectorize(lambda x: isalive(x, pat_except=self.pat_except))(
            self.meas
        )

        msg = "前年まで生存。枯死？"
        errors = [
            ErrDat(self.plot_id, self.rec_id[i], self.col_meas[j], msg)
            for i, j in zip(*np.where(match_na))
            if j > 0 and alive[i, j - 1]
        ]
        return errors

    def check_values_after_d(self):
        """
        Check if the value after 'd' is correct (na or dd).

        dの次の値がnaあるいはddになっていない
        """
        pat_d = r"^d$"
        pat_dd = r"^dd$|^na$|^NA$"
        match_d = np.vectorize(lambda x: find_pattern(x, pat_d))(self.meas)
        match_dd = np.vectorize(lambda x: find_pattern(x, pat_dd))(self.meas)

        msg = "枯死の次の調査時のgbhが「na」もしくは「dd」になっていない"
        errors = [
            ErrDat(self.plot_id, self.rec_id[i], self.col_meas[j], msg)
            for i, j in zip(*np.where(match_d))
            if j < (len(self.col_meas) - 1) and not match_dd[i, j + 1]
        ]
        return errors

    def find_anomaly(self):
        """
        Find anomaly in GBH growth.

        成長量が基準より大きいあるいは小さい
        """
        errors = []
        if self.meas.shape[1] == 1:
            return errors

        # NOTE: 前回の値にcd, vn, viが付く場合はスキップ
        meas_c = np.vectorize(lambda x: isvalid(x, return_value=True))(self.meas)
        pat_vc = r"^vi|^vn|^cd"
        match_vc = np.vectorize(lambda x: find_pattern(x, pat_vc))(self.meas)

        for i, row in enumerate(meas_c):
            index_notnull = np.where(~np.isnan(row))[0]
            if len(index_notnull) == 0:
                continue
            gbhdiff = np.diff(row[index_notnull])
            yrdiff = np.diff(np.vectorize(retrive_year)(self.col_meas[index_notnull]))

            excess = gbhdiff > yrdiff * 2.5 + 3.8
            minus = gbhdiff < -3.1
            index_excess = index_notnull[1:][excess]
            index_minus = index_notnull[1:][minus]

            msg = "成長量が基準値より大きい。測定ミス？"
            errors.extend(
                [
                    ErrDat(self.plot_id, self.rec_id[i], self.col_meas[j], msg)
                    for j in index_excess
                    if not match_vc[i, j - 1]
                ]
            )

            msg = "成長量が基準値より小さい。測定ミス？"
            errors.extend(
                [
                    ErrDat(self.plot_id, self.rec_id[i], self.col_meas[j], msg)
                    for j in index_minus
                    if not match_vc[i, j - 1]
                ]
            )
        return errors

    def check_values_recruits(self):
        """
        Check if the size of the recruit is too large when it is first recorded.

        新規加入個体（naの次のgbhが数値）のサイズが基準より大きい
        """
        errors = []
        if self.meas.shape[1] > 1:
            return errors

        meas_c = np.vectorize(lambda x: isvalid(x, return_value=True))(self.meas)
        notnull = ~np.isnan(meas_c)
        pat_na = r"^na$|^NA$"
        match_na = np.vectorize(lambda x: find_pattern(x, pat_na))(self.meas)
        msg = "新規加入個体だが、加入時のgbhが基準より大きいのため前回計測忘れの疑い"
        errors = []
        for i, j in zip(*np.where(match_na[:, :-1] & notnull[:, 1:])):
            yrdiff = np.diff(np.vectorize(retrive_year)(self.col_meas[[j, j + 1]]))[0]
            if meas_c[i, j + 1] >= (15 + yrdiff * 2.5 + 3.8):
                target = "{}={}; {}={}".format(
                    self.col_meas[j],
                    self.meas_orig[i, j],
                    self.col_meas[j + 1],
                    self.meas_orig[i, j + 1],
                )
                errors.append(ErrDat(self.plot_id, self.rec_id[i], target, msg))
        return errors

    def check_values_nd(self):
        """
        Check if the input value is marked with "nd", but is within normal growth range.

        ndだが前後の測定値と比較して成長量の基準に収まっている
        """
        pat_ndxx = r"^nd\s?([0-9]+[.]?[0-9]*)"
        match_ndxx = np.vectorize(lambda x: find_pattern(x, pat_ndxx))(self.meas)
        meas_c = np.vectorize(lambda x: isvalid(x, "^nd", return_value=True))(self.meas)

        errors = []
        for i, row in enumerate(meas_c):
            if any(match_ndxx[i]):
                index_notnull = np.where(~np.isnan(row))[0]
                if len(index_notnull) == 0:
                    continue
                gbhdiff = np.diff(row[index_notnull])
                yrdiff = np.diff(
                    np.vectorize(retrive_year)(self.col_meas[index_notnull])
                )
                in_range = (gbhdiff <= yrdiff * 2.5 + 3.8) & (gbhdiff >= -3.1)
                index_notnull[1:][in_range]
                msg = "誤って「nd: 測定間違い」となっている可能性あり"
                for jj in np.where(match_ndxx[i][index_notnull][1:])[0]:
                    j = index_notnull[jj + 1]
                    # 最初か最後の調査時のndはスキップ
                    if (j < (len(self.col_meas) - 1)) and (jj < (len(in_range) - 1)):
                        if all(in_range[jj : (jj + 2)]):
                            errors.append(
                                ErrDat(
                                    self.plot_id, self.rec_id[i], self.col_meas[j], msg
                                )
                            )
        return errors

    def check_all(self, throughly=False):
        """
        Run data checks.

        すべての項目をチェック
        """
        errors = []
        errors.extend(self.check_invalid_date())
        errors.extend(self.check_sp_not_in_list())
        errors.extend(self.check_synonym())
        errors.extend(self.check_tag_dup())
        errors.extend(self.check_indv_null())
        errors.extend(self.check_sp_mismatch())
        errors.extend(self.check_mesh_xy())
        errors.extend(self.check_stem_xy())
        errors.extend(self.check_blank_in_data_cols())
        errors.extend(self.check_invalid_values())
        self.mask_invalid_values()
        self.replace_dxx_in_gbh()
        errors.extend(self.check_missing())
        errors.extend(self.check_values_after_d())
        errors.extend(self.find_anomaly())
        errors.extend(self.check_values_recruits())
        errors.extend(self.check_values_nd())
        if throughly:
            errors.extend(self.check_local_name())
        return errors


class CheckDataLitter(CheckDataCommon):
    """
    Check litter data.

    リターデータのチェック
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.period = np.array(
            [i + "-" + j for i, j in self.select(["s_date1", "s_date2"])]
        )

    def check_trap_date_combinations(self):
        """
        Check for redundant or missing traps in the same installation period.

        同じ設置・回収日でトラップの重複・欠落がないか
        """
        # 同じ時期でも数日に分けて回収した場合などもあり、必ずしもエラーではない
        errors = []
        trap_in_use = [k for k, v in self.trap_list.items() if v["use"]]
        for period_s, n_trap in zip(*np.unique(self.period, return_counts=True)):
            s_date1_s, s_date2_s = period_s.split("-")
            trap_s = self.select("trap_id")[np.where(self.period == period_s)[0]]
            trap_dup = find_duplicates(trap_s)
            if trap_dup.size > 0:
                msg = "同じ設置・回収日の組み合わせでトラップの重複あり"
                msg += " ({})".format("; ".join(trap_dup))
                errors.append(ErrDat(self.plot_id, s_date1_s, "", msg))

            if n_trap < len(trap_in_use):
                msg = "同じ設置・回収日の組み合わせでトラップの欠落あり"
                trap_lack = [x for x in trap_in_use if x not in trap_s]
                msg += " ({})".format(";".join(trap_lack))
                errors.append(ErrDat(self.plot_id, s_date1_s, "", msg))
        return errors

    def check_installation_period1(self):
        """
        Check if the installation period is too long or too short.

        設置期間が長い/短い
        """
        errors = []
        s_date1_s, s_date2_s = zip(*[p.split("-") for p in np.unique(self.period)])
        s_date1_s = np.array(s_date1_s)
        s_date2_s = np.array(s_date2_s)
        s_date1_s_dt = np.array(list(map(as_datetime, s_date1_s)))
        s_date2_s_dt = np.array(list(map(as_datetime, s_date2_s)))
        delta_days = np.vectorize(calc_delta_days)(s_date1_s_dt, s_date2_s_dt)

        long_d = delta_days > 45
        # 越冬設置は除外
        overwinter_site = [
            "UR-BC1",
            "AS-DB1",
            "AS-DB2",
            "TM-DB1",
            "OY-DB1",
            "KY-DB1",
            "OT-EC1",
            "OG-DB1",
        ]
        within_year = np.array(
            [i.year == j.year for i, j in zip(s_date1_s_dt, s_date2_s_dt)]
        )
        msg = "設置期間が46日以上"
        if self.plot_id in overwinter_site:
            errors.extend(
                [
                    ErrDat(self.plot_id, d1, "", msg)
                    for d1 in s_date1_s[long_d & within_year]
                ]
            )
        else:
            errors.extend(
                [ErrDat(self.plot_id, d1, "", msg) for d1 in s_date1_s[long_d]]
            )

        # 設置期間が短い
        short_d = delta_days < 11
        msg = "設置期間が10日以下"
        errors.extend([ErrDat(self.plot_id, d1, "", msg) for d1 in s_date1_s[short_d]])

        return errors

    def check_installation_period2(self):
        """
        Check if installation periods vary by trap.

        設置期間がトラップによって異なる
        """
        s_date1_s, s_date2_s = zip(*[p.split("-") for p in np.unique(self.period)])
        s_date1_s = np.array(s_date1_s)
        s_date2_s = np.array(s_date2_s)
        msg = "設置期間がトラップによって異なる"
        errors = [
            ErrDat(self.plot_id, d1, "", msg)
            for d1 in np.unique(s_date1_s)
            if len(s_date2_s[s_date1_s == d1]) > 1
        ]
        return errors

    def check_installation_period3(self):
        """
        Check mismatch between installation date and previous collection date.

        設置日と前回の回収日のずれ
        """
        errors = []
        for trap in np.unique(self.trap_id):
            s_date1_s = self.select("s_date1")[self.trap_id == trap]
            s_date2_s = self.select("s_date2")[self.trap_id == trap]
            s_date1_s_dt = np.array(list(map(as_datetime, s_date1_s)))
            s_date2_s_dt = np.array(list(map(as_datetime, s_date2_s)))

            delta_days = np.vectorize(calc_delta_days)(
                s_date2_s_dt[:-1], s_date1_s_dt[1:]
            )

            interrupted = (delta_days != 0) & (delta_days < 45)
            within_year = np.array(
                [i.year == j.year for i, j in zip(s_date2_s_dt[:-1], s_date1_s_dt[1:])]
            )

            msg = "前回の回収日から{}日間の中断期間"
            errors.extend(
                [
                    ErrDat(
                        self.plot_id, s_date1_s[i + 1], trap, msg.format(delta_days[i])
                    )
                    for i in np.where(interrupted & within_year)[0]
                ]
            )
        return errors

    def find_anomaly(self):
        """
        Find anomaly in litter weigtht measurements.

        重量データの異常値の検出
        """
        # 器官・回収日ごとにスミルノフ-グラブス検定により外れ値を検出
        # NOTE: -> Tukey's fences（箱ひげ図の外れ値検出法, ノンパラ）に変更
        # 数値以外の文字列はnanに置換
        # 0が多い月はそれ以外の値が外れ値になるので、0も除外
        # 絶乾重量（wdry）のみ
        wdry_cols = [i for i, x in enumerate(self.col_meas) if re.search("wdry_", x)]
        meas_wdry = self.meas[:, wdry_cols]
        meas_wdry[meas_wdry == "0"] = np.nan
        meas_c = np.vectorize(lambda x: isvalid(x, "^NA$|^na$|^-$", return_value=True))(
            meas_wdry
        )
        meas_c[np.less(meas_c, 0.0, where=~np.isnan(meas_c))] = np.nan

        msg = "{}は外れ値の可能性あり"
        errors = []
        for period_s in np.unique(self.period):
            d1 = period_s.split("-")[0]
            values_s = meas_c[self.period == period_s]
            trap_id_s = self.trap_id[self.period == period_s]

            for j, vals in enumerate(np.transpose(values_s)):
                if np.sum(~np.isnan(vals)) < 5:
                    continue
                # outlier = smirnov_grubbs(np.log(vals), alpha=0.01)
                outlier = find_anomaly_tukey(np.log(vals), k=3)
                errors.extend(
                    [
                        ErrDat(
                            self.plot_id,
                            d1,
                            trap_id_s[i],
                            msg.format(self.col_meas[wdry_cols[j]]),
                        )
                        for i in np.where(outlier)[0]
                    ]
                )
        return errors

    def check_all(self, throughly=False):
        """
        Run data checks.

        すべての項目をチェック
        """
        errors = []
        errors.extend(self.check_invalid_date())
        # 日付に不正な入力値がある場合はここで終了
        if errors:
            logger.warning("日付に不正な入力値があるためデータチェックを終了")
            return errors
        errors.extend(self.check_trap_date_combinations())
        errors.extend(self.check_installation_period1())
        errors.extend(self.check_installation_period2())
        errors.extend(self.check_installation_period3())
        errors.extend(self.check_blank_in_data_cols())
        errors.extend(self.check_invalid_values())
        self.mask_invalid_values()
        errors.extend(self.check_positive())
        errors.extend(self.find_anomaly())

        return errors


class CheckDataSeed(CheckDataCommon):
    """
    Check seed data.

    種子データのチェック
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.period = np.array(
            [i + "-" + j for i, j in self.select(["s_date1", "s_date2"])]
        )

    def check_trap(self):
        """
        Check for consistency with the trap list.

        トラップリストとの整合性チェック
        """
        errors = []
        if not self.trap_list:
            return errors
        trap_uniq = np.unique(self.trap_id)
        trap_not_in_list = [i for i in trap_uniq if i not in self.trap_list]
        for trap in trap_not_in_list:
            msg = "リストにないtrap_id ({})".format(trap)
            errors.extend(
                [
                    ErrDat(self.plot_id, self.rec_id[i], trap, msg)
                    for i in np.where(self.trap_id == trap)[0]
                ]
            )
        return errors

    def check_all(self, throughly=False):
        """
        Run data checks.

        すべての項目をチェック
        """
        errors = []
        errors.extend(self.check_invalid_date())
        errors.extend(self.check_sp_not_in_list())
        errors.extend(self.check_synonym())
        errors.extend(self.check_blank_in_data_cols())
        errors.extend(self.check_trap())
        errors.extend(self.check_invalid_values())
        errors.extend(self.check_positive())
        if throughly:
            errors.extend(self.check_local_name())

        return errors


def isvalid(s: str, pat_except="", return_value=False):
    """Check if the value is a numeric or one of the exceptions."""
    r = re.compile(pat_except)
    try:
        s = r.sub("", str(s))
        f = float(np.nan if s == "" else s)
        return f if return_value else True
    except ValueError:
        return np.nan if return_value else False


def find_pattern(s: str, pat: str):
    """Find a pattern in the given string and return a Boolean value."""
    # if isinstance(pat, str):
    r = re.compile(pattern=pat)
    if r.match(str(s)):
        return True
    else:
        return False


def isalive(s: str, pat_except: str = "", gbh_threthold: float = 15.0):
    """
    Check whether the GBH larger than the threthold.

    Parameters
    ----------
    s : str
        input
    pat_except : list
        Regular expression pattern of exception strings
    gbh_threthold : float, default 15.0
        GBH threthod

    """
    x = isvalid(s, pat_except, return_value=True)
    if not np.isnan(x) and (x >= gbh_threthold):
        return True
    else:
        return False


def isdate(s_date: str, if_nan=False):
    """Check whether the date string is yyyymmdd format or not."""
    if len(s_date) != 8:
        return False
    try:
        datetime.strptime(s_date, "%Y%m%d")
        return True
    except ValueError:
        return False
    except TypeError:
        return if_nan


def as_datetime(s_date, format="%Y%m%d"):
    """Convert a string in yyyymmdd format to a datetime object."""
    if isdate(s_date, if_nan=False):
        return datetime.strptime(s_date, format)
    else:
        # return np.datetime64("NaT")
        return None


def find_duplicates(array):
    """Find duplicates."""
    uniq, counts = np.unique(array, return_counts=True)
    return uniq[counts > 1]


def dt64_to_dt(dt64):
    """Convert the numpy.datetime64 object to the datetime object."""
    try:
        return datetime.strptime(dt64.astype("datetime64[D]").astype("str"), "%Y-%m-%d")
    except ValueError:
        return np.nan


def return_growth_year(dt):
    """
    Return the growth year.

    調査日から成長年を返す（8月以前に調査した場合は前年を成長年とする）
    """
    if isinstance(dt, np.datetime64):
        dt = dt64_to_dt(dt)
    try:
        if dt.month < 8:
            return dt.year - 1
        else:
            return dt.year
    except AttributeError:
        return -1


def retrive_year(x: str) -> int:
    """
    Retrive the year from the string.

    Parameters
    ----------
    x : str
        String contains year (e.g. "gbh04", where "04" means 2004)

    """
    r = re.compile(r"[0-9]+")
    errmsg = "Can not retrive a year from {!r}".format(x)
    match = r.search(x)
    if match:
        if len(match.group()) == 2:
            form = "%y"
        elif len(match.group()) == 4:
            form = "%Y"
        else:
            raise RuntimeError(errmsg)
        return int(datetime.strptime(match.group(), form).strftime("%Y"))
    else:
        raise RuntimeError(errmsg)


def calc_delta_days(d1, d2):
    """Calculate the day difference between datetime objects."""
    if isinstance(d1, datetime) and isinstance(d2, datetime):
        return (d2 - d1).days
    else:
        raise TypeError("d1 and d2 must be datetime.datetime objects")


def find_anomaly_tukey(x, k=3.0):
    """
    Outlier detection by using the Tukey's fences.

    Parameters
    ----------
    x : numpy array or list
        Input array
    k : float, default 3.0
        Non-negative constant to determine threatholds for outlier detection.
        k = 1.5 indicates an "outlier" (corresponding to standard boxplot),
        and k = 3 indicates data that is "far out" (default 3)

    """
    x = np.array(x)
    mask = ~np.isnan(x)
    out = mask.copy()
    xs = x[mask]
    q1, q3 = np.percentile(xs, q=[25, 75])
    iqr = q3 - q1
    lower_bound = q1 - k * iqr
    upper_bound = q3 + k * iqr
    out[np.where(mask)] = np.where((xs < lower_bound) | (xs > upper_bound), True, False)
    return out


# def smirnov_grubbs(x, alpha=0.05):
#     """Outlier detection by Smirnov-Grubbs test."""
#     import scipy.stats as stats
#
#     x = np.array(x)
#     xs = x[~np.isnan(x)]
#     i_out = np.array([], dtype=np.int64)
#
#     while True:
#         n = len(xs)
#         if n < 5:
#             break
#         t = stats.t.isf(q=((alpha / n) / 2), df=(n - 2))
#         tau = (n - 1) * t / np.sqrt(n * (n - 2) + n * t**2)
#         mu, sd = xs.mean(), xs.std(ddof=1)
#         if sd == 0:
#             break
#         far = xs.max() if np.abs(xs.max() - mu) > np.abs(xs.min() - mu) else xs.min()
#         tau_far = np.abs((far - mu) / sd)
#         if tau_far < tau:
#             break
#         i_out = np.append(i_out, np.where(x == far)[0])
#         xs = np.delete(xs, np.where(xs == far)[0])
#
#     return np.array([True if i in i_out else False for i in range(len(x))])


def argsort_n(x: np.ndarray) -> List[int]:
    """Return the indices that would sort an array in a natural sort order."""

    def key(string):
        return [
            int(s) if s.isdigit() else s.lower() for s in re.split("([0-9]+)", string)
        ]

    return sorted(range(len(x)), key=lambda i: key(str(x[i])))


def sort_array(x: np.ndarray, sort_col: Union[int, List[int]] = []) -> np.ndarray:
    """
    Sort a two dimentional array by column values.

    Parameters
    ----------
    x : numpy array
        Two dimentional array to be sorted
    sort_col : int or List[int]
        Column indices to use for sorting. All columns will be used in default.

    """
    if not sort_col:
        sort_col = list(range(x.shape[1]))[::-1]
    elif not isinstance(sort_col, list):
        sort_col = [sort_col]
    try:
        for c in sort_col:
            x = x[argsort_n(x[:, c])]
    except IndexError as e:
        if re.search("out of bounds", e.__str__()):
            raise
        else:
            raise TypeError("sort_col must be an int or List[int]")
    else:
        return x


def save_errors_to_xlsx(
    errors: List[ErrDat], dest_filepath: str, header: Optional[List[str]] = None
):
    """
    Save the error list in a xlsx file.

    Parameters
    ----------
    errors : List[ErrDat]
        List of ErrDat objects
    dest_filepath : str
        Destination file path where you want to save an error list
    header : List[str], optional
        Column headers

    """
    errors_a = np.array([astuple(e) for e in errors])
    errors_a = sort_array(errors_a)

    wb = Workbook()
    # ws = wb.create_sheet("確認事項{}".format(datetime.now().strftime("%y%m%d")))
    ws = wb.active
    ws.title = "確認事項{}".format(datetime.now().strftime("%y%m%d"))

    start_row = 1
    if header:
        if "サイトでの対応" not in header:
            header.append("サイトでの対応")
        for j, value in enumerate(header):
            ws.cell(start_row, 1 + j, value)
            ws.cell(start_row, 1 + j).font = Font(name="Arial", bold=True)
            ws.cell(start_row, 1 + j).fill = PatternFill("solid", fgColor="7e7e7e")
        start_row += 1

    for i, row in enumerate(errors_a):
        for j, x in enumerate(row):
            ws.cell(start_row + i, 1 + j, x)
            ws.cell(start_row + i, 1 + j).font = Font(name="Arial")

    wb.save(dest_filepath)
    wb.close()


def check_data(
    d: Optional[MonitoringData] = None,
    filepath: Optional[str] = None,
    path_spdict: Optional[str] = None,
    path_xy: Optional[str] = None,
    path_trap: Optional[str] = None,
    path_ignore: Optional[str] = None,
    throughly: bool = False,
):
    """
    Check errors in Moni-Sen data.

    Parameters
    ----------
    d : MonitoringData object, optional
        Data to be checked
    filepath : str, optional
        Path to the data file
    path_spdict : str
        Path to the file of the Japanese species name dictionary
    path_xy : str
        Path to the file of xy coordinates of 10 x 10 m grids in the plot
    path_trap : str
        Path to the file of a trap list
    path_ignore : str
        Path to file of an ignore list for data checking
    throughly : bool, default False
        If True, all checking tasks are executed.

    """
    if not d and filepath:
        d = read_data(filepath)
    elif not d:
        raise RuntimeError("'d' or 'filepath' is needed")

    cd: Union[CheckDataTree, CheckDataLitter, CheckDataSeed]
    if d.data_type == "treeGBH":
        cd = CheckDataTree(**vars(d), path_spdict=path_spdict, path_xy=path_xy)
    elif d.data_type == "litter":
        cd = CheckDataLitter(**vars(d), path_trap=path_trap)
    elif d.data_type == "seed":
        cd = CheckDataSeed(**vars(d), path_spdict=path_spdict, path_trap=path_trap)
    else:
        raise TypeError("'data_type' does not defined")

    errors = cd.check_all(throughly=throughly)

    if path_ignore and errors:
        # 無視リストにあるエラー項目を除外
        ignore = [ErrDat(*i) for i in read_table(path_ignore)[:, :4]]
        errors = [x for x in errors if x not in ignore]

    return errors
